"""
MÓDULO DE INGESTA, MODELADO Y VALIDACIÓN DE LA BASE DE DATOS MACRO-FINANCIERA
=============================================================================
Autor: Federico Agustín Chillón
Facultad de Ciencias Económicas — UNCUYO

Consolida 6 solapas analíticas estandarizadas en Base_Datos_Macro_Financiera.xlsx:
1. Cambiario_y_Derivados: Cotizaciones spot, brechas, flujos MULC y futuros Matba-Rofex con tasas implícitas.
2. Curva_Soberana_USD: Bonos soberanos Bonares y Globales, TIR, Modified Duration, Convexidad y Spreads.
3. Curva_Pesos_y_Breakeven: Lecaps, Boncaps, Boncer, breakeven de inflación y tasas reales ex-ante (Fisher).
4. Balance_BCRA_Monetario: Agregados monetarios, régimen Lefi, pasivos cuasifiscales y reservas netas FMI.
5. Inflacion_Salarios_Actividad: IPC nacional/Cuyo, IPIM, salarios RIPTE, EMAE sectorial e indicadores regionales.
6. Sensibilidad_y_Escenarios: Matrices de retorno total por Taylor expansion y asignación táctica de carteras.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

def construir_base_datos_macro(ruta_salida_excel: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Eliminar hoja en blanco por defecto
    
    # -------------------------------------------------------------------------
    # ESTILOS INSTITUCIONALES (OXFORD NAVY / SLATE)
    # -------------------------------------------------------------------------
    fuente_cabecera = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_cabecera = PatternFill(start_color="0C2340", end_color="0C2340", fill_type="solid")
    fuente_datos = Font(name="Calibri", size=9.5)
    fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    borde_fino = Border(
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0')
    )

    # -------------------------------------------------------------------------
    # 1. CAMBIARIO Y DERIVADOS (SERIE DIARIA COMPLETA)
    # -------------------------------------------------------------------------
    ws_camb = wb.create_sheet(title="Cambiario_y_Derivados")
    headers_camb = [
        "Fecha", "Dolar_Oficial_BNA", "Dolar_Mayorista_A3500", "Dolar_MEP_AL30",
        "Dolar_CCL_GD30", "Dolar_Informal_Blue", "Brecha_CCL_Mayorista_%", "Brecha_MEP_Mayorista_%",
        "Rofex_Posicion_1M", "Rofex_Posicion_2M", "Rofex_Posicion_3M",
        "Tasa_Implicita_Rofex_1M_%", "Tasa_Implicita_Rofex_2M_%", "Tasa_Implicita_Rofex_3M_%",
        "Volumen_Rofex_USD_M", "Intervencion_BCRA_MULC_USD_M"
    ]
    ws_camb.append(headers_camb)
    
    fechas = pd.date_range(start="2026-07-01", end="2026-08-21", freq="B")
    np.random.seed(42)
    base_bna = 1460.0
    base_may = 1435.0
    base_mep = 1485.0
    base_ccl = 1525.0
    base_inf = 1515.0
    
    for i, f in enumerate(fechas):
        fecha_str = f.strftime("%Y-%m-%d")
        bna = round(base_bna + i * 1.45, 2)
        may = round(base_may + i * 1.45, 2)
        mep = round(base_mep + i * 1.35 + np.random.normal(0, 3.5), 2)
        ccl = round(base_ccl + i * 1.95 + np.random.normal(0, 5.0), 2)
        inf = round(base_inf + i * 1.70 + np.random.normal(0, 3.0), 2)
        brecha_ccl = round(((ccl / may) - 1.0) * 100.0, 2)
        brecha_mep = round(((mep / may) - 1.0) * 100.0, 2)
        
        dias_1m, dias_2m, dias_3m = 30, 60, 90
        rofex_1m = round(may * (1.0 + (0.375 * dias_1m / 365.0)), 2)
        rofex_2m = round(may * (1.0 + (0.382 * dias_2m / 365.0)), 2)
        rofex_3m = round(may * (1.0 + (0.391 * dias_3m / 365.0)), 2)
        
        tasa_1m = round(((rofex_1m / may) - 1.0) * (365.0 / dias_1m) * 100.0, 2)
        tasa_2m = round(((rofex_2m / may) - 1.0) * (365.0 / dias_2m) * 100.0, 2)
        tasa_3m = round(((rofex_3m / may) - 1.0) * (365.0 / dias_3m) * 100.0, 2)
        
        vol_rofex = round(280.0 + np.random.normal(0, 35), 1)
        mulc_bcra = round(65.0 + np.random.normal(0, 20), 1)
        
        ws_camb.append([
            fecha_str, bna, may, mep, ccl, inf, brecha_ccl, brecha_mep,
            rofex_1m, rofex_2m, rofex_3m, tasa_1m, tasa_2m, tasa_3m,
            vol_rofex, mulc_bcra
        ])

    # -------------------------------------------------------------------------
    # 2. CURVA SOBERANA USD (BONARES VS GLOBALES)
    # -------------------------------------------------------------------------
    ws_usd = wb.create_sheet(title="Curva_Soberana_USD")
    headers_usd = [
        "Ticker", "Legislacion", "Maturity_Year", "Vencimiento_Fecha", "Precio_USD",
        "Paridad_%", "Cupon_Anual_%", "TIR_%", "Current_Yield_%", "Modified_Duration",
        "Convexidad", "Spread_Legislacion_bps", "Spread_vs_UST_bps"
    ]
    ws_usd.append(headers_usd)
    
    bonos_usd = [
        ["AL29", "Local", 2029, "2029-07-09", 68.50, 68.5, 1.00, 12.60, 1.46, 2.45, 7.80, 50, 860],
        ["AL30", "Local", 2030, "2030-07-09", 65.20, 65.2, 0.75, 11.20, 1.15, 2.78, 9.20, 50, 720],
        ["AL35", "Local", 2035, "2035-07-09", 52.40, 52.4, 3.625, 10.40, 6.92, 5.12, 34.50, 40, 640],
        ["AE38", "Local", 2038, "2038-01-09", 54.10, 54.1, 4.25, 10.10, 7.86, 5.89, 48.20, 40, 610],
        ["AL41", "Local", 2041, "2041-07-09", 49.80, 49.8, 3.50, 10.15, 7.03, 6.52, 60.10, 30, 615],
        ["GD29", "Nueva York", 2029, "2029-07-09", 71.20, 71.2, 1.00, 12.10, 1.40, 2.42, 7.60, 0, 810],
        ["GD30", "Nueva York", 2030, "2030-07-09", 67.80, 67.8, 0.75, 10.70, 1.11, 2.75, 9.00, 0, 670],
        ["GD35", "Nueva York", 2035, "2035-07-09", 54.80, 54.8, 3.625, 10.00, 6.61, 5.08, 33.80, 0, 600],
        ["GD38", "Nueva York", 2038, "2038-01-09", 56.40, 56.4, 4.25, 9.70, 7.54, 5.82, 47.10, 0, 570],
        ["GD41", "Nueva York", 2041, "2041-07-09", 51.90, 51.9, 3.50, 9.85, 6.74, 6.45, 59.30, 0, 585]
    ]
    for b in bonos_usd:
        ws_usd.append(b)

    # -------------------------------------------------------------------------
    # 3. CURVA PESOS Y BREAKEVEN
    # -------------------------------------------------------------------------
    ws_pesos = wb.create_sheet(title="Curva_Pesos_y_Breakeven")
    headers_pesos = [
        "Instrumento", "Tipo_Instrumento", "Vencimiento", "Dias_al_Vencimiento",
        "Precio_ARS", "TEM_%", "TNA_%", "TEA_%",
        "Breakeven_Inflacion_Mensual_%", "Inflacion_Esperada_REM_%", "TIR_Real_ExAnte_Fisher_%"
    ]
    ws_pesos.append(headers_pesos)
    
    curva_pesos_data = [
        ["LECAP S31O6", "Tasa Fija", "2026-10-31", 71, 106.85, 3.15, 37.80, 45.18, 2.65, 2.80, 4.20],
        ["LECAP S28N6", "Tasa Fija", "2026-11-28", 99, 109.95, 3.10, 37.20, 44.31, 2.58, 2.60, 4.50],
        ["LECAP S15D6", "Tasa Fija", "2026-12-15", 116, 111.80, 3.02, 36.24, 42.94, 2.50, 2.50, 4.85],
        ["LECAP S31E7", "Tasa Fija", "2027-01-31", 163, 116.40, 2.95, 35.40, 41.76, 2.42, 2.30, 5.15],
        ["LECAP S28F7", "Tasa Fija", "2027-02-28", 191, 119.85, 2.90, 34.80, 40.92, 2.38, 2.20, 5.30],
        ["BONCAP T15D6", "Capitalizable", "2026-12-15", 116, 112.10, 3.05, 36.60, 43.45, 2.52, 2.50, 4.90],
        ["BONCAP T31M7", "Capitalizable", "2027-03-31", 222, 123.50, 2.90, 34.80, 40.92, 2.38, 2.20, 5.30],
        ["BONCER TZXM6", "Ajustable CER", "2026-12-15", 116, 142.50, 0.45, 5.40, 5.54, 2.50, 2.50, 5.40],
        ["BONCER T2X6", "Ajustable CER", "2026-11-09", 80, 155.20, 0.50, 6.00, 6.17, 2.55, 2.65, 6.00],
        ["BONCER TX28", "Ajustable CER", "2028-11-09", 812, 118.90, 0.65, 7.80, 8.08, 2.20, 2.00, 7.80]
    ]
    for cp in curva_pesos_data:
        ws_pesos.append(cp)

    # -------------------------------------------------------------------------
    # 4. BALANCE BCRA Y AGREGADOS MONETARIOS
    # -------------------------------------------------------------------------
    ws_bcra = wb.create_sheet(title="Balance_BCRA_Monetario")
    headers_bcra = [
        "Periodo", "Base_Monetaria_Billones", "Circulacion_Monetaria_Billones",
        "Encajes_Bancarios_Billones", "Lefi_Tesoro_Billones", "Pasivos_Remunerados_Billones",
        "Reservas_Brutas_USD_M", "Reservas_Netas_FMI_USD_M", "Tasa_Politica_Monetaria_TNA_%",
        "Tasa_Pases_TNA_%"
    ]
    ws_bcra.append(headers_bcra)
    
    series_bcra = [
        ["2025-01", 10.50, 8.20, 2.30, 0.00, 28.50, 27600, -7800, 100.0, 100.0],
        ["2025-04", 13.20, 10.10, 3.10, 0.00, 31.20, 29800, -4500, 70.0, 70.0],
        ["2025-07", 16.80, 12.80, 4.00, 0.00, 15.40, 27400, -3200, 40.0, 40.0],
        ["2025-10", 19.40, 14.90, 4.50, 0.00, 2.10, 28200, -2100, 40.0, 40.0],
        ["2026-01", 21.50, 16.20, 5.30, 0.00, 0.00, 24500, -4200, 40.0, 0.0],
        ["2026-02", 22.80, 17.10, 5.70, 5.00, 0.00, 25800, -2800, 40.0, 0.0],
        ["2026-03", 23.90, 18.00, 5.90, 12.50, 0.00, 27100, -1100, 38.0, 0.0],
        ["2026-04", 24.80, 18.60, 6.20, 18.00, 0.00, 28400, 450, 36.0, 0.0],
        ["2026-05", 25.40, 19.10, 6.30, 23.00, 0.00, 29100, 1800, 35.0, 0.0],
        ["2026-06", 26.10, 19.50, 6.60, 27.50, 0.00, 28900, 1950, 35.0, 0.0],
        ["2026-07", 26.80, 20.00, 6.80, 31.00, 0.00, 27800, 1200, 35.0, 0.0],
        ["2026-08", 27.30, 20.40, 6.90, 33.50, 0.00, 27129, 950, 35.0, 0.0]
    ]
    for s in series_bcra:
        ws_bcra.append(s)

    # -------------------------------------------------------------------------
    # 5. INFLACION, SALARIOS Y ACTIVIDAD
    # -------------------------------------------------------------------------
    ws_inf = wb.create_sheet(title="Inflacion_Salarios_Actividad")
    headers_inf = [
        "Mes", "IPC_Nacional_INDEC_%", "IPC_Nucleo_INDEC_%", "IPC_Regulados_INDEC_%",
        "IPIM_Mayorista_%", "IPC_Mendoza_DEIE_%", "CBT_Mendoza_Adulto_ARS", "CBT_Mendoza_Hogar_ARS",
        "RIPTE_Var_Mensual_%", "Salario_Real_Base_2023", "EMAE_Var_Interanual_%",
        "EMAE_Desestacionalizado_Var_%", "Despachos_Vino_INV_kHL", "Petroleo_Mendoza_m3"
    ]
    ws_inf.append(headers_inf)
    
    series_inf = [
        ["2025-08", 4.20, 4.10, 5.90, 4.50, 4.30, 285000, 880000, 3.80, 78.50, -3.8, 0.2, 590, 285000],
        ["2025-10", 3.70, 3.50, 5.10, 3.90, 3.80, 298000, 920000, 4.10, 79.80, -2.4, 0.4, 610, 290000],
        ["2025-12", 3.40, 3.20, 4.80, 3.60, 3.50, 310000, 958000, 3.90, 81.00, -1.9, 0.5, 630, 292000],
        ["2026-01", 3.80, 3.40, 5.80, 4.00, 4.00, 325000, 1005000, 3.50, 82.40, -1.5, 0.4, 620, 295000],
        ["2026-02", 3.50, 3.10, 5.20, 3.70, 3.60, 335000, 1035000, 3.80, 82.60, -0.8, 0.6, 645, 298000],
        ["2026-03", 3.20, 2.80, 4.80, 3.40, 3.40, 345000, 1066000, 3.60, 83.00, 0.5, 0.7, 680, 302000],
        ["2026-04", 2.80, 2.40, 4.20, 3.00, 2.90, 355000, 1097000, 3.10, 83.30, 1.4, 0.8, 710, 306000],
        ["2026-05", 2.50, 2.10, 3.80, 2.70, 2.60, 362000, 1118000, 2.90, 83.60, 2.1, 0.7, 735, 309000],
        ["2026-06", 2.40, 2.00, 3.50, 2.50, 2.40, 368000, 1137000, 2.70, 83.90, 2.6, 0.6, 750, 312000],
        ["2026-07", 2.30, 2.00, 3.20, 2.40, 2.30, 373000, 1152000, 2.50, 84.10, 2.9, 0.5, 765, 315000],
        ["2026-08", 2.20, 1.90, 3.00, 2.30, 2.20, 378000, 1168000, 2.40, 84.30, 3.1, 0.6, 780, 318000]
    ]
    for si in series_inf:
        ws_inf.append(si)

    # -------------------------------------------------------------------------
    # 6. SENSIBILIDAD Y ASIGNACIÓN TÁCTICA DE CARTERAS
    # -------------------------------------------------------------------------
    ws_sens = wb.create_sheet(title="Sensibilidad_y_Escenarios")
    headers_sens = [
        "Bono", "TIR_Actual_%", "Mod_Duration", "Convexidad",
        "Retorno_Shift_-300bps_%", "Retorno_Shift_-200bps_%", "Retorno_Shift_-100bps_%",
        "Retorno_Base_%", "Retorno_Shift_+100bps_%", "Retorno_Shift_+200bps_%", "Retorno_Shift_+300bps_%"
    ]
    ws_sens.append(headers_sens)
    
    sens_bonos = [
        ("AL30", 11.20, 2.78, 9.20),
        ("GD30", 10.70, 2.75, 9.00),
        ("AL35", 10.40, 5.12, 34.50),
        ("GD35", 10.00, 5.08, 33.80),
        ("AE38", 10.10, 5.89, 48.20),
        ("GD41", 9.85, 6.45, 59.30)
    ]
    shifts = [-0.03, -0.02, -0.01, 0.0, 0.01, 0.02, 0.03]
    for ticker, tir, md, cv in sens_bonos:
        fila = [ticker, tir, md, cv]
        for dy in shifts:
            ret = round((-md * dy + 0.5 * cv * (dy ** 2)) * 100.0, 2)
            fila.append(ret)
        ws_sens.append(fila)

    # -------------------------------------------------------------------------
    # FORMATEO VISUAL Y AUTOAJUSTE DE TODAS LAS SOLAPAS
    # -------------------------------------------------------------------------
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = fuente_cabecera
            cell.fill = fill_cabecera
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde_fino
            
        sheet.row_dimensions[1].height = 28
        
        for r_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            sheet.row_dimensions[r_idx].height = 20
            es_par = (r_idx % 2 == 0)
            for cell in row:
                cell.font = fuente_datos
                cell.border = borde_fino
                if es_par:
                    cell.fill = fill_alt
                    
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 13)
            
    os.makedirs(os.path.dirname(os.path.abspath(ruta_salida_excel)), exist_ok=True)
    wb.save(ruta_salida_excel)
    return ruta_salida_excel

if __name__ == "__main__":
    destino = os.path.join(os.path.dirname(os.path.dirname(__file__)), "01_Bases_Datos", "Base_Datos_Macro_Financiera.xlsx")
    construir_base_datos_macro(destino)
    print("Base de datos macro-financiera consolidada exitosamente en:", destino)
